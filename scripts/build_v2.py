# -*- coding: utf-8 -*-
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule

HERE = os.path.dirname(os.path.abspath(__file__))
FONT_NAME = "Arial"

BLUE = Font(name=FONT_NAME, color="0000FF")
BLACK = Font(name=FONT_NAME, color="000000")
BOLD = Font(name=FONT_NAME, bold=True)
BOLD_WHITE = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
ITALIC_GRAY = Font(name=FONT_NAME, italic=True, color="808080")
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="203864")
COMPANY_FILL = PatternFill("solid", fgColor="4472C4")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0;(#,##0);-'

with open(os.path.join(HERE, "dart_data2.json"), encoding="utf-8") as f:
    DATA = json.load(f)["companies"]

COMPANIES = ["한국전력", "LG에너지솔루션", "HD현대"]
YEAR_COLS = ["B", "C", "D"]
PERIOD_KEYS = ["bfefrmtrm", "frmtrm", "thstrm"]

METRICS = [
    ("revenue", "매출액", False),
    ("cogs", "매출원가", False),
    ("gross_profit", "매출총이익", True),
    ("op_income", "영업이익", False),
    ("net_income", "당기순이익", False),
    ("curr_assets", "유동자산", False),
    ("inventory", "재고자산", False),
    ("curr_liab", "유동부채", False),
    ("total_assets", "자산총계", False),
    ("total_liab", "부채총계", False),
    ("total_equity", "자본총계", False),
    ("cash_dividend_total", "현금배당금총액", False),
]

wb = Workbook()

# =========================================================================
# 사용안내
# =========================================================================
ws0 = wb.active
ws0.title = "사용안내"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 100
ws0["A1"] = "3개 기업 재무제표 비교분석기 (DART Open API 연동)"
ws0["A1"].font = TITLE_FONT

lines = [
    "",
    "1. 데이터 출처",
    "   - 금융감독원 전자공시시스템(DART) Open API (opendart.fss.or.kr)",
    "   - 각 기업 최신 사업보고서(연간) 기준 연결재무제표(CFS, 연결 재무제표 미제출 시 별도(OFS))를 사용했습니다.",
    "   - 배당 정보는 DART '배당에 관한 사항(alotMatter)' 공시를 사용했으며, 원 단위(백만원 -> 원 환산)로 표시했습니다.",
    "   - '원본데이터' 시트 각 회사 블록 상단에 회사별 corp_code, 종목코드, 사업연도, 재무제표 구분을 표기했습니다.",
    "",
    "2. 이 파일의 구성",
    "   - 원본데이터 : DART에서 가져온 3개 기업 x 3개년 원본 재무제표/배당 수치",
    "   - 비교분석   : 원본데이터를 참조하는 수식으로 수익성/안정성/성장성/배당 지표를 자동 계산, 3개 기업을 나란히 비교",
    "   - 대시보드   : 매출액·영업이익·순이익 추이 및 수익성/안정성/배당성향 비교 차트, 최신연도 요약 스코어카드",
    "",
    "3. 주의사항",
    "   - 이 데이터는 스크립트 실행 시점(DART 최신 공시 기준)의 스냅샷입니다. 최신 공시가 갱신되면 다시 조회해야 합니다.",
    "   - 계정과목명은 기업/업종에 따라 다르게 표기될 수 있어 자동 매칭을 사용했습니다.",
    "   - 배당을 실시하지 않았거나 공시에 해당 항목이 없는 경우 '-' 로 표시됩니다 (예: 성장 단계 기업의 무배당).",
    "   - 값은 원(KRW) 단위입니다.",
]
for i, line in enumerate(lines, start=2):
    c = ws0.cell(row=i, column=1, value=line)
    c.font = BOLD if (line and not line.startswith(" ")) else BLACK

# =========================================================================
# 원본데이터
# =========================================================================
ws1 = wb.create_sheet("원본데이터")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 20
for col in YEAR_COLS:
    ws1.column_dimensions[col].width = 20

ws1.merge_cells("A1:D1")
ws1["A1"] = "3개 기업 재무제표 원본 데이터 (단위: 원)"
ws1["A1"].font = TITLE_FONT

row_map = {}
r = 3
for company in COMPANIES:
    info = DATA.get(company, {})
    if "error" in info:
        ws1.cell(row=r, column=1, value=f"{company}: 데이터 조회 실패 ({info['error']})").font = Font(name=FONT_NAME, color="FF0000", bold=True)
        r += 2
        continue

    years = [info["bsns_year"] - 2, info["bsns_year"] - 1, info["bsns_year"]]

    ws1.merge_cells(f"A{r}:D{r}")
    c = ws1.cell(row=r, column=1, value=f"■ {company} ({info['dart_corp_name']})")
    c.font = BOLD_WHITE
    c.fill = COMPANY_FILL
    r += 1

    source_note = (f"출처: DART Open API, corp_code={info['corp_code']}, 종목코드={info['stock_code']}, "
                   f"사업연도={info['bsns_year']}(사업보고서), 재무제표구분={info['fs_div']} "
                   f"({'연결' if info['fs_div'] == 'CFS' else '별도'}), 배당: alotMatter 공시")
    ws1.merge_cells(f"A{r}:D{r}")
    ws1.cell(row=r, column=1, value=source_note).font = ITALIC_GRAY
    r += 1

    col_header_row = r
    ws1.cell(row=r, column=1, value="항목").font = BOLD_WHITE
    ws1.cell(row=r, column=1).fill = HEADER_FILL
    for col, yr in zip(YEAR_COLS, years):
        cell = ws1[f"{col}{r}"]
        cell.value = str(yr)
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    r += 1

    metric_rows = {"header": col_header_row, "years": years}
    for key, label, is_formula in METRICS:
        ws1.cell(row=r, column=1, value=label).font = BLACK
        if not is_formula:
            if key == "cash_dividend_total":
                periods = info.get("dividend", {}).get("cash_dividend_total", {})
            else:
                periods = info[key]
            for col, pk in zip(YEAR_COLS, PERIOD_KEYS):
                val = periods.get(pk)
                cell = ws1[f"{col}{r}"]
                cell.value = val if val is not None else "N/A"
                cell.font = BLUE
                cell.number_format = NUM_FMT
                cell.border = BORDER
        metric_rows[key] = r
        r += 1

    gp_row = metric_rows["gross_profit"]
    rev_row = metric_rows["revenue"]
    cogs_row = metric_rows["cogs"]
    for col in YEAR_COLS:
        cell = ws1[f"{col}{gp_row}"]
        cell.value = f"=IFERROR({col}{rev_row}-{col}{cogs_row},\"N/A\")"
        cell.font = BLACK
        cell.number_format = NUM_FMT
        cell.border = BORDER

    row_map[company] = metric_rows
    r += 1

ws1.freeze_panes = "B4"

with open(os.path.join(HERE, "row_map2.json"), "w", encoding="utf-8") as f:
    json.dump(row_map, f, ensure_ascii=False, indent=2)

wb.save(os.path.join(HERE, "재무비교분석기2_stage1.xlsx"))
print("stage1 v2 saved")
