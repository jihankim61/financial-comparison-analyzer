# -*- coding: utf-8 -*-
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

HERE = os.path.dirname(os.path.abspath(__file__))
FONT_NAME = "Arial"

BLUE = Font(name=FONT_NAME, color="0000FF")
BLACK = Font(name=FONT_NAME, color="000000")
BOLD = Font(name=FONT_NAME, bold=True)
BOLD_WHITE = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
ITALIC_GRAY = Font(name=FONT_NAME, italic=True, color="808080")
RED_BOLD = Font(name=FONT_NAME, bold=True, color="C00000")
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="203864")
COMPANY_FILL = PatternFill("solid", fgColor="4472C4")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0;(#,##0);-'

with open(os.path.join(HERE, "dart_data4.json"), encoding="utf-8") as f:
    DATA = json.load(f)["companies"]

COMPANIES = ["한국전력", "삼성전자", "SK하이닉스"]
YEAR_COLS = ["B", "C", "D"]
PERIOD_KEYS = ["bfefrmtrm", "frmtrm", "thstrm"]

METRICS = [
    ("revenue", "매출액", False), ("cogs", "매출원가", False), ("gross_profit", "매출총이익", True),
    ("op_income", "영업이익", False), ("net_income", "당기순이익", False),
    ("curr_assets", "유동자산", False), ("inventory", "재고자산", False), ("curr_liab", "유동부채", False),
    ("total_assets", "자산총계", False), ("total_liab", "부채총계", False), ("total_equity", "자본총계", False),
]

wb = Workbook()

# ============ 사용안내 ============
ws0 = wb.active
ws0.title = "사용안내"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 104
ws0["A1"] = "3개 기업 재무제표 비교분석기 + 배당소득 분리과세 대상 판정"
ws0["A1"].font = TITLE_FONT
lines = [
    "",
    "1. 데이터 출처",
    "   - 금융감독원 전자공시시스템(DART) Open API (opendart.fss.or.kr)",
    "   - 재무제표: 각 기업 최신 사업보고서(연간) 기준 연결재무제표(CFS, 미제출 시 별도 OFS)",
    "   - 배당: DART '배당에 관한 사항(alotMatter)' 공시. 현금배당금총액과 기업이 직접 공시한 (연결)현금배당성향(%)을 함께 사용",
    "",
    "2. 이 파일의 구성",
    "   - 원본데이터   : 3개 기업 x 3개년 원본 재무제표 수치",
    "   - 비교분석     : 수익성/안정성/성장성 지표 자동 계산 비교",
    "   - 대시보드     : 추이 차트 및 최신연도 요약 스코어카드",
    "   - 분리과세판정 : 2026년 시행 예정 배당소득 분리과세 특례의 '고배당기업' 요건 충족 여부 추정",
    "",
    "3. 분리과세판정 시트 관련 중요 안내",
    "   - 이 시트는 조세특례제한법 제104조의27(고배당기업 주식 배당소득에 대한 과세특례, 2025.12.23. 신설,",
    "     국가법령정보센터 확인)의 요건을 그대로 반영했습니다. 다만 시행령 세부사항(공시방법 등)은 반영하지 못했습니다.",
    "   - 반드시 국세청(nts.go.kr)·기획재정부 공식 자료 또는 세무전문가 확인 후 실제 세무 판단에 활용하시기 바랍니다.",
    "   - 판정 대상은 '기업의 법인세'가 아니라, 그 기업으로부터 배당을 받는 개인 주주가 종합과세 대신",
    "     분리과세를 신청할 수 있는지 여부입니다.",
]
for i, line in enumerate(lines, start=2):
    c = ws0.cell(row=i, column=1, value=line)
    c.font = BOLD if (line and not line.startswith(" ")) else BLACK

# ============ 원본데이터 ============
ws1 = wb.create_sheet("원본데이터")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 20
for col in YEAR_COLS:
    ws1.column_dimensions[col].width = 20
ws1.merge_cells("A1:D1")
ws1["A1"] = "3개 기업 재무제표 원본 데이터 (단위: 원)"
ws1["A1"].font = TITLE_FONT

row_map = {}
r = 3
for company in COMPANIES:
    info = DATA[company]
    years = [info["bsns_year"] - 2, info["bsns_year"] - 1, info["bsns_year"]]
    ws1.merge_cells(f"A{r}:D{r}")
    c = ws1.cell(row=r, column=1, value=f"■ {company} ({info['dart_corp_name']})")
    c.font = BOLD_WHITE
    c.fill = COMPANY_FILL
    r += 1
    source_note = (f"출처: DART Open API, corp_code={info['corp_code']}, 종목코드={info['stock_code']}, "
                   f"사업연도={info['bsns_year']}(사업보고서), 재무제표구분={info['fs_div']}"
                   f"({'연결' if info['fs_div'] == 'CFS' else '별도'})")
    ws1.merge_cells(f"A{r}:D{r}")
    ws1.cell(row=r, column=1, value=source_note).font = ITALIC_GRAY
    r += 1
    col_header_row = r
    ws1.cell(row=r, column=1, value="항목").font = BOLD_WHITE
    ws1.cell(row=r, column=1).fill = HEADER_FILL
    for col, yr in zip(YEAR_COLS, years):
        cell = ws1[f"{col}{r}"]
        cell.value = str(yr)
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    r += 1
    metric_rows = {"header": col_header_row, "years": years}
    for key, label, is_formula in METRICS:
        ws1.cell(row=r, column=1, value=label).font = BLACK
        if not is_formula:
            periods = info[key]
            for col, pk in zip(YEAR_COLS, PERIOD_KEYS):
                val = periods.get(pk)
                cell = ws1[f"{col}{r}"]
                cell.value = val if val is not None else "N/A"
                cell.font = BLUE
                cell.number_format = NUM_FMT
                cell.border = BORDER
        metric_rows[key] = r
        r += 1
    gp_row, rev_row, cogs_row = metric_rows["gross_profit"], metric_rows["revenue"], metric_rows["cogs"]
    for col in YEAR_COLS:
        cell = ws1[f"{col}{gp_row}"]
        cell.value = f"=IFERROR({col}{rev_row}-{col}{cogs_row},\"N/A\")"
        cell.font = BLACK
        cell.number_format = NUM_FMT
        cell.border = BORDER
    row_map[company] = metric_rows
    r += 1
ws1.freeze_panes = "B4"

# ============ 비교분석 ============
ws2 = wb.create_sheet("비교분석")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 16
for col in YEAR_COLS:
    ws2.column_dimensions[col].width = 14
ws2.merge_cells("A1:D1")
ws2["A1"] = "3개 기업 재무비율 비교분석"
ws2["A1"].font = TITLE_FONT
YEARS = row_map[COMPANIES[0]]["years"]
RAW = "원본데이터"

RATIO_SECTIONS = [
    ("[ 수익성 지표 ]", [
        ("매출총이익률", "gross_profit", "revenue", "0.0%"),
        ("영업이익률", "op_income", "revenue", "0.0%"),
        ("순이익률", "net_income", "revenue", "0.0%"),
        ("ROA (총자산순이익률)", "net_income", "total_assets", "0.0%"),
        ("ROE (자기자본순이익률)", "net_income", "total_equity", "0.0%"),
    ]),
    ("[ 안정성 지표 ]", [
        ("부채비율", "total_liab", "total_equity", "0.0%"),
        ("유동비율", "curr_assets", "curr_liab", "0.0%"),
        ("자기자본비율", "total_equity", "total_assets", "0.0%"),
    ]),
]
ratio_row_map = {}

def write_ratio_table(start_row, section_title, ratio_defs):
    r = start_row
    ws2.merge_cells(f"A{r}:D{r}")
    ws2.cell(row=r, column=1, value=section_title).font = BOLD
    ws2.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    for label, num_key, den_key, fmt in ratio_defs:
        ws2.cell(row=r, column=1, value=label).font = BOLD_WHITE
        ws2.cell(row=r, column=1).fill = HEADER_FILL
        for col, yr in zip(YEAR_COLS, YEARS):
            c = ws2[f"{col}{r}"]
            c.value = str(yr)
            c.font = BOLD_WHITE
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        r += 1
        ratio_row_map[label] = {}
        for company in COMPANIES:
            rm = row_map[company]
            ws2.cell(row=r, column=1, value=company).font = BLACK
            for col in YEAR_COLS:
                cell = ws2[f"{col}{r}"]
                cell.value = f"=IFERROR('{RAW}'!{col}{rm[num_key]}/'{RAW}'!{col}{rm[den_key]},\"-\")"
                cell.number_format = fmt
                cell.font = BLACK
                cell.border = BORDER
            ratio_row_map[label][company] = r
            r += 1
        r += 1
    return r

row_cursor = 3
for section_title, defs in RATIO_SECTIONS:
    row_cursor = write_ratio_table(row_cursor, section_title, defs)

GROWTH_DEFS = [("매출액증가율", "revenue"), ("영업이익증가율", "op_income"),
               ("순이익증가율", "net_income"), ("총자산증가율", "total_assets")]
r = row_cursor
ws2.merge_cells(f"A{r}:D{r}")
ws2.cell(row=r, column=1, value="[ 성장성 지표 (전년 대비 증가율) ]").font = BOLD
ws2.cell(row=r, column=1).fill = SECTION_FILL
r += 1
for label, key in GROWTH_DEFS:
    ws2.cell(row=r, column=1, value=label).font = BOLD_WHITE
    ws2.cell(row=r, column=1).fill = HEADER_FILL
    for col, yr in zip(YEAR_COLS, YEARS):
        c = ws2[f"{col}{r}"]
        c.value = str(yr)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    r += 1
    ratio_row_map[label] = {}
    for company in COMPANIES:
        rm = row_map[company]
        ws2.cell(row=r, column=1, value=company).font = BLACK
        ws2[f"B{r}"] = "N/A"
        ws2[f"B{r}"].alignment = Alignment(horizontal="center")
        ws2[f"C{r}"] = f"=IFERROR(('{RAW}'!C{rm[key]}-'{RAW}'!B{rm[key]})/'{RAW}'!B{rm[key]},\"-\")"
        ws2[f"D{r}"] = f"=IFERROR(('{RAW}'!D{rm[key]}-'{RAW}'!C{rm[key]})/'{RAW}'!C{rm[key]},\"-\")"
        for col in ["C", "D"]:
            ws2[f"{col}{r}"].number_format = "0.0%"
            ws2[f"{col}{r}"].font = BLACK
        for col in YEAR_COLS:
            ws2[f"{col}{r}"].border = BORDER
        ratio_row_map[label][company] = r
        r += 1
    r += 1
ws2.freeze_panes = "B3"

# ============ 대시보드 ============
ws3 = wb.create_sheet("대시보드")
ws3.sheet_view.showGridLines = False
for col, w in zip(["A", "B", "C", "D", "E", "F", "G", "H"], [16, 12, 12, 12, 4, 16, 12, 12]):
    ws3.column_dimensions[col].width = w
ws3.merge_cells("A1:H1")
ws3["A1"] = "3개 기업 재무비교 대시보드 (한국전력 / 삼성전자 / SK하이닉스)"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A2:H2")
ws3["A2"] = f"기준: {YEARS[0]}~{YEARS[-1]} 사업보고서 (연결재무제표 기준, DART Open API)"
ws3["A2"].font = Font(name=FONT_NAME, italic=True, color="808080")

def add_line_chart(anchor, title, metric_key):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height = 8
    chart.width = 16
    cat_row = row_map[COMPANIES[0]]["header"]
    cats = Reference(ws1, min_col=2, max_col=4, min_row=cat_row, max_row=cat_row)
    for company in COMPANIES:
        row = row_map[company][metric_key]
        data = Reference(ws1, min_col=2, max_col=4, min_row=row, max_row=row)
        chart.series.append(Series(data, title=company))
    chart.set_categories(cats)
    for s in chart.series:
        s.smooth = False
        s.marker.symbol = "circle"
    ws3.add_chart(chart, anchor)

def add_ratio_bar_chart(anchor, title, ratio_label):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.style = 10
    chart.height = 8
    chart.width = 16
    rows = ratio_row_map[ratio_label]
    header_row = list(rows.values())[0] - 1
    cats = Reference(ws2, min_col=2, max_col=4, min_row=header_row, max_row=header_row)
    for company in COMPANIES:
        row = rows[company]
        data = Reference(ws2, min_col=2, max_col=4, min_row=row, max_row=row)
        chart.series.append(Series(data, title=company))
    chart.set_categories(cats)
    ws3.add_chart(chart, anchor)

add_line_chart("A4", "매출액 추이 (원)", "revenue")
add_line_chart("A21", "영업이익 추이 (원)", "op_income")
add_line_chart("A38", "당기순이익 추이 (원)", "net_income")
add_ratio_bar_chart("J4", "ROE 비교 (%)", "ROE (자기자본순이익률)")
add_ratio_bar_chart("J21", "부채비율 비교 (%)", "부채비율")
add_ratio_bar_chart("J38", "영업이익률 비교 (%)", "영업이익률")

score_start = 55
ws3.merge_cells(f"A{score_start}:H{score_start}")
ws3.cell(row=score_start, column=1, value=f"■ {YEARS[-1]}년 요약 스코어카드").font = BOLD
ws3.cell(row=score_start, column=1).fill = SECTION_FILL
SCORE_METRICS = [("영업이익률", "0.0%"), ("순이익률", "0.0%"), ("ROE (자기자본순이익률)", "0.0%"),
                 ("부채비율", "0.0%"), ("유동비율", "0.0%"), ("매출액증가율", "0.0%")]
hdr_row = score_start + 1
ws3.cell(row=hdr_row, column=1, value="회사명").font = BOLD_WHITE
ws3.cell(row=hdr_row, column=1).fill = HEADER_FILL
for j, (label, fmt) in enumerate(SCORE_METRICS):
    cell = ws3.cell(row=hdr_row, column=2 + j, value=label)
    cell.font = BOLD_WHITE
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
ws3.row_dimensions[hdr_row].height = 30
first_data_row = hdr_row + 1
for i, company in enumerate(COMPANIES):
    r = first_data_row + i
    ws3.cell(row=r, column=1, value=company).font = BLACK
    for j, (label, fmt) in enumerate(SCORE_METRICS):
        src_row = ratio_row_map[label][company]
        cell = ws3.cell(row=r, column=2 + j)
        cell.value = f"='비교분석'!D{src_row}"
        cell.number_format = fmt
        cell.font = BLACK
        cell.border = BORDER
last_data_row = first_data_row + len(COMPANIES) - 1
for j in range(len(SCORE_METRICS)):
    col_letter = chr(ord("B") + j)
    rng = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
    rule = ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50,
                          mid_color="FFEB84", end_type="max", end_color="63BE7B")
    ws3.conditional_formatting.add(rng, rule)

# ============ 분리과세판정 ============
ws4 = wb.create_sheet("분리과세판정")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 34
for col in ["B", "C", "D", "E"]:
    ws4.column_dimensions[col].width = 16

ws4.merge_cells("A1:E1")
ws4["A1"] = "배당소득 분리과세 대상(고배당기업) 요건 판정"
ws4["A1"].font = TITLE_FONT

ws4.merge_cells("A2:E2")
ws4["A2"] = "조세특례제한법 제104조의27(고배당기업 주식 배당소득에 대한 과세특례, 2025.12.23. 신설) · 2026.1.1. 이후 지급 배당분부터 2028년까지 한시 적용"
ws4["A2"].font = Font(name=FONT_NAME, italic=True, color="808080")

disclaimer_lines = [
    "⚠ 주의: 이 판정은 조세특례제한법 제104조의27 조문(국가법령정보센터, 2025.12.23. 신설)을 근거로 구성했으나, 시행령 세부사항은 반영하지 못했습니다.",
    "  - 요건(법조문): ① 직전 사업연도 배당소득이 2024사업연도보다 비감소, AND ② 직전 사업연도 배당성향 40% 이상 OR (배당성향 25% 이상 AND 전전사업연도 대비 배당금 10% 이상 증가)",
    "  - 배당성향은 각 기업이 DART에 직접 공시한 (연결)현금배당성향(%)을 사용합니다 (원본데이터 시트의 자체 계산 순이익 기준과 소수 오차 있을 수 있음).",
    "  - 코넥스 상장·투자회사 여부, 펀드/리츠/SPC 등 간접투자 제외 요건, 기업의 자체 요건 공시 여부는 확인하지 않았습니다.",
    "  - 세율: 2천만원 이하 14%, 2천만원초과~3억원 20%(누진), 3억원초과~50억원 25%(누진), 50억원초과 30%(누진) — 초안 발표 당시(35% 3단계)와 실제 신설 조문의 세율이 다르므로 확정 조문 기준으로 표기.",
    "  - 최종 판단은 국세청·기획재정부 공식 발표 또는 세무전문가 확인 후 하시기 바랍니다.",
]
r = 3
for line in disclaimer_lines:
    ws4.merge_cells(f"A{r}:E{r}")
    cell = ws4.cell(row=r, column=1, value=line)
    cell.font = RED_BOLD if r == 3 else Font(name=FONT_NAME, size=10.5, color="7F6000")
    cell.fill = WARN_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[r].height = 16
    r += 1
r += 1

tax_years = ["2022", "2023", "2024", "2025"]
tax_cols = ["B", "C", "D", "E"]

verdict_summary_row = r
ws4.merge_cells(f"A{r}:E{r}")
ws4.cell(row=r, column=1, value="■ 요약 판정 결과 (2025 사업연도 기준)").font = BOLD
ws4.cell(row=r, column=1).fill = SECTION_FILL
r += 2

summary_rows = {}
for company in COMPANIES:
    summary_rows[company] = r
    ws4.cell(row=r, column=1, value=company).font = BOLD
    r += 1
r += 1

company_blocks = {}
for company in COMPANIES:
    info = DATA[company]
    ws4.merge_cells(f"A{r}:E{r}")
    c = ws4.cell(row=r, column=1, value=f"■ {company} ({info['dart_corp_name']})")
    c.font = BOLD_WHITE
    c.fill = COMPANY_FILL
    r += 1

    rcept = info["dividend_source_rcept_no"]
    note = f"출처: DART alotMatter 공시 (접수번호 {rcept.get('report_a (year, year-1, year-2)')}, {rcept.get('report_b (year-1, year-2, year-3)')})"
    ws4.merge_cells(f"A{r}:E{r}")
    ws4.cell(row=r, column=1, value=note).font = ITALIC_GRAY
    r += 1

    hdr_row_local = r
    ws4.cell(row=r, column=1, value="항목").font = BOLD_WHITE
    ws4.cell(row=r, column=1).fill = HEADER_FILL
    for col, yr in zip(tax_cols, tax_years):
        cell = ws4[f"{col}{r}"]
        cell.value = yr
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    r += 1

    div_row = r
    ws4.cell(row=r, column=1, value="현금배당금총액(원)").font = BLACK
    for col, yr in zip(tax_cols, tax_years):
        val = info["dividend_krw_by_year"].get(yr)
        cell = ws4[f"{col}{r}"]
        cell.value = val if val is not None else "N/A"
        cell.font = BLUE
        cell.number_format = NUM_FMT
        cell.border = BORDER
    r += 1

    pay_row = r
    ws4.cell(row=r, column=1, value="(연결)현금배당성향 [DART 공시]").font = BLACK
    for col, yr in zip(tax_cols, tax_years):
        val = info["dart_reported_payout_pct_by_year"].get(yr)
        cell = ws4[f"{col}{r}"]
        cell.value = (val / 100) if val is not None else "N/A"
        cell.font = BLUE
        cell.number_format = "0.0%"
        cell.border = BORDER
    r += 1

    condA_row = r
    ws4.cell(row=r, column=1, value="① 직전사업연도 배당, 2024사업연도 대비 비감소").font = BLACK
    ws4[f"E{r}"] = (f'=IF(OR(NOT(ISNUMBER(E{div_row})),NOT(ISNUMBER(D{div_row}))),"판단불가",'
                     f'IF(E{div_row}>=D{div_row},"충족","미충족"))')
    ws4[f"E{r}"].font = BLACK
    ws4[f"E{r}"].border = BORDER
    r += 1

    growth_row = r
    ws4.cell(row=r, column=1, value="전전사업연도 대비 증가율").font = BLACK
    ws4[f"E{r}"] = f'=IFERROR(E{div_row}/D{div_row}-1,"-")'
    ws4[f"E{r}"].font = BLACK
    ws4[f"E{r}"].number_format = "0.0%"
    ws4[f"E{r}"].border = BORDER
    r += 1

    condB_row = r
    ws4.cell(row=r, column=1, value="② 배당성향 40%+ 또는 (25%+ 및 전전사업연도 대비 10%+ 증가)").font = BLACK
    ws4[f"E{r}"] = (f'=IF(NOT(ISNUMBER(E{pay_row})),"판단불가",'
                     f'IF(E{pay_row}>=0.4,"충족(40%+)",'
                     f'IF(AND(E{pay_row}>=0.25,ISNUMBER(E{growth_row}),E{growth_row}>=0.1),'
                     f'"충족(25%+ 및 전전사업연도대비10%+증가)","미충족")))')
    ws4[f"E{r}"].font = BLACK
    ws4[f"E{r}"].border = BORDER
    r += 1

    final_row = r
    ws4.cell(row=r, column=1, value="▶ 최종 판정 (①AND②, 추정)").font = BOLD
    ws4[f"E{r}"] = (f'=IF(OR(E{condA_row}="판단불가",E{condB_row}="판단불가"),"판단보류(자료부족, 공시 확인 필요)",'
                     f'IF(AND(E{condA_row}="충족",LEFT(E{condB_row},2)="충족"),'
                     f'"분리과세 대상 요건 충족 (추정)","요건 미충족 (추정)"))')
    ws4[f"E{r}"].font = BOLD
    ws4[f"E{r}"].border = BORDER
    r += 2

    company_blocks[company] = {"final_row": final_row}
    # link into summary
    ws4[f"B{summary_rows[company]}"] = f"=E{final_row}"
    ws4[f"B{summary_rows[company]}"].font = BOLD
    ws4.merge_cells(f"B{summary_rows[company]}:E{summary_rows[company]}")

# conditional formatting for verdict cells (summary + per-block final rows)
verdict_ranges = [f"B{summary_rows[c]}" for c in COMPANIES] + [f"E{company_blocks[c]['final_row']}" for c in COMPANIES]
for rng in verdict_ranges:
    ws4.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("대상 요건 충족",{rng}))'], fill=GREEN_FILL))
    ws4.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("판단보류",{rng}))'], fill=AMBER_FILL))
    ws4.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("요건 미충족",{rng}))'], fill=RED_FILL))

wb.save(os.path.join(HERE, "재무비교분석기4_분리과세.xlsx"))

with open(os.path.join(HERE, "row_map4.json"), "w", encoding="utf-8") as f:
    json.dump(row_map, f, ensure_ascii=False, indent=2)
with open(os.path.join(HERE, "ratio_row_map4.json"), "w", encoding="utf-8") as f:
    json.dump(ratio_row_map, f, ensure_ascii=False, indent=2)

print("v4 workbook (with tax eligibility) saved")
