# -*- coding: utf-8 -*-
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule

HERE = os.path.dirname(os.path.abspath(__file__))
FONT_NAME = "Arial"

BLACK = Font(name=FONT_NAME, color="000000")
BOLD = Font(name=FONT_NAME, bold=True)
BOLD_WHITE = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="203864")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

with open(os.path.join(HERE, "row_map2.json"), encoding="utf-8") as f:
    ROW_MAP = json.load(f)
with open(os.path.join(HERE, "ratio_row_map2.json"), encoding="utf-8") as f:
    RATIO_ROW_MAP = json.load(f)

COMPANIES = ["한국전력", "LG에너지솔루션", "HD현대"]
YEARS = ROW_MAP[COMPANIES[0]]["years"]
RAW = "원본데이터"
CMP = "비교분석"

wb = load_workbook(os.path.join(HERE, "재무비교분석기2_stage2.xlsx"))
ws_raw = wb[RAW]
ws_cmp = wb[CMP]

ws3 = wb.create_sheet("대시보드")
ws3.sheet_view.showGridLines = False
for col, w in zip(["A", "B", "C", "D", "E", "F", "G", "H"], [16, 12, 12, 12, 4, 16, 12, 12]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells("A1:H1")
ws3["A1"] = "3개 기업 재무비교 대시보드 (한국전력 / LG에너지솔루션 / HD현대)"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A2:H2")
ws3["A2"] = f"기준: {YEARS[0]}~{YEARS[-1]} 사업보고서 (연결재무제표 기준, DART Open API)"
ws3["A2"].font = Font(name=FONT_NAME, italic=True, color="808080")


def add_line_chart(anchor, title, metric_key):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height = 8
    chart.width = 16
    cat_row = ROW_MAP[COMPANIES[0]]["header"]
    cats = Reference(ws_raw, min_col=2, max_col=4, min_row=cat_row, max_row=cat_row)
    for company in COMPANIES:
        row = ROW_MAP[company][metric_key]
        data = Reference(ws_raw, min_col=2, max_col=4, min_row=row, max_row=row)
        series = Series(data, title=company)
        chart.series.append(series)
    chart.set_categories(cats)
    for s in chart.series:
        s.smooth = False
        s.marker.symbol = "circle"
    ws3.add_chart(chart, anchor)


def add_ratio_bar_chart(anchor, title, ratio_label):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.style = 10
    chart.height = 8
    chart.width = 16
    rows = RATIO_ROW_MAP[ratio_label]
    header_row = list(rows.values())[0] - 1
    cats = Reference(ws_cmp, min_col=2, max_col=4, min_row=header_row, max_row=header_row)
    for company in COMPANIES:
        row = rows[company]
        data = Reference(ws_cmp, min_col=2, max_col=4, min_row=row, max_row=row)
        series = Series(data, title=company)
        chart.series.append(series)
    chart.set_categories(cats)
    ws3.add_chart(chart, anchor)


add_line_chart("A4", "매출액 추이 (원)", "revenue")
add_line_chart("A21", "영업이익 추이 (원)", "op_income")
add_line_chart("A38", "당기순이익 추이 (원)", "net_income")

add_ratio_bar_chart("J4", "ROE 비교 (%)", "ROE (자기자본순이익률)")
add_ratio_bar_chart("J21", "부채비율 비교 (%)", "부채비율")
add_ratio_bar_chart("J38", "배당성향 비교 (%)", "배당성향")

# ---------------------------------------------------------------
# Summary scorecard for latest year
# ---------------------------------------------------------------
score_start = 55
ws3.merge_cells(f"A{score_start}:H{score_start}")
ws3.cell(row=score_start, column=1, value=f"■ {YEARS[-1]}년 요약 스코어카드").font = BOLD
ws3.cell(row=score_start, column=1).fill = SECTION_FILL

SCORE_METRICS = [
    ("영업이익률", "0.0%"),
    ("순이익률", "0.0%"),
    ("ROE (자기자본순이익률)", "0.0%"),
    ("부채비율", "0.0%"),
    ("유동비율", "0.0%"),
    ("배당성향", "0.0%"),
    ("매출액증가율", "0.0%"),
]

hdr_row = score_start + 1
ws3.cell(row=hdr_row, column=1, value="회사명").font = BOLD_WHITE
ws3.cell(row=hdr_row, column=1).fill = HEADER_FILL
for j, (label, fmt) in enumerate(SCORE_METRICS):
    cell = ws3.cell(row=hdr_row, column=2 + j, value=label)
    cell.font = BOLD_WHITE
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
ws3.row_dimensions[hdr_row].height = 30

first_data_row = hdr_row + 1
for i, company in enumerate(COMPANIES):
    r = first_data_row + i
    ws3.cell(row=r, column=1, value=company).font = BLACK
    for j, (label, fmt) in enumerate(SCORE_METRICS):
        src_row = RATIO_ROW_MAP[label][company]
        cell = ws3.cell(row=r, column=2 + j)
        cell.value = f"='{CMP}'!D{src_row}"
        cell.number_format = fmt
        cell.font = BLACK
        cell.border = BORDER

last_data_row = first_data_row + len(COMPANIES) - 1
for j in range(len(SCORE_METRICS)):
    col_letter = chr(ord("B") + j)
    rng = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws3.conditional_formatting.add(rng, rule)

note_row = last_data_row + 2
ws3.merge_cells(f"A{note_row}:H{note_row}")
ws3.cell(row=note_row, column=1,
         value="※ 배당성향이 '-'로 표시된 경우 해당 연도 배당을 실시하지 않았거나 DART 공시에 항목이 없는 경우입니다 (예: LG에너지솔루션).").font = Font(name=FONT_NAME, italic=True, color="808080")

wb.save(os.path.join(HERE, "재무비교분석기2.xlsx"))
print("stage3 v2 (final) saved")
