# -*- coding: utf-8 -*-
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule

HERE = os.path.dirname(os.path.abspath(__file__))
FONT_NAME = "Arial"

BLUE = Font(name=FONT_NAME, color="0000FF")
BLACK = Font(name=FONT_NAME, color="000000")
BOLD = Font(name=FONT_NAME, bold=True)
BOLD_WHITE = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
ITALIC_GRAY = Font(name=FONT_NAME, italic=True, color="808080")
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="203864")
COMPANY_FILL = PatternFill("solid", fgColor="4472C4")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0;(#,##0);-'

with open(os.path.join(HERE, "dart_data3.json"), encoding="utf-8") as f:
    DATA = json.load(f)["companies"]

COMPANIES = ["LG유플러스", "HD현대", "셀트리온"]
YEAR_COLS = ["B", "C", "D"]
PERIOD_KEYS = ["bfefrmtrm", "frmtrm", "thstrm"]

METRICS = [
    ("revenue", "매출액", False),
    ("cogs", "매출원가", False),
    ("gross_profit", "매출총이익", True),
    ("op_income", "영업이익", False),
    ("net_income", "당기순이익", False),
    ("curr_assets", "유동자산", False),
    ("inventory", "재고자산", False),
    ("curr_liab", "유동부채", False),
    ("total_assets", "자산총계", False),
    ("total_liab", "부채총계", False),
    ("total_equity", "자본총계", False),
    ("cash_dividend_total", "현금배당금총액", False),
]

wb = Workbook()

# ============ 사용안내 ============
ws0 = wb.active
ws0.title = "사용안내"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 100
ws0["A1"] = "3개 기업 재무제표 비교분석기 (DART Open API 연동)"
ws0["A1"].font = TITLE_FONT
lines = [
    "",
    "1. 데이터 출처",
    "   - 금융감독원 전자공시시스템(DART) Open API (opendart.fss.or.kr)",
    "   - 각 기업 최신 사업보고서(연간) 기준 연결재무제표(CFS, 연결 재무제표 미제출 시 별도(OFS))를 사용했습니다.",
    "   - 배당 정보는 DART '배당에 관한 사항(alotMatter)' 공시를 사용했으며, 원 단위(백만원 -> 원 환산)로 표시했습니다.",
    "   - '원본데이터' 시트 각 회사 블록 상단에 회사별 corp_code, 종목코드, 사업연도, 재무제표 구분을 표기했습니다.",
    "",
    "2. 이 파일의 구성",
    "   - 원본데이터 : DART에서 가져온 3개 기업 x 3개년 원본 재무제표/배당 수치",
    "   - 비교분석   : 원본데이터를 참조하는 수식으로 수익성/안정성/성장성/배당 지표를 자동 계산, 3개 기업을 나란히 비교",
    "   - 대시보드   : 매출액·영업이익·순이익 추이 및 수익성/안정성/배당성향 비교 차트, 최신연도 요약 스코어카드",
    "",
    "3. 주의사항",
    "   - 이 데이터는 스크립트 실행 시점(DART 최신 공시 기준)의 스냅샷입니다. 최신 공시가 갱신되면 다시 조회해야 합니다.",
    "   - LG유플러스는 통신업 특성상 '매출원가'를 별도 계정으로 공시하지 않아 매출총이익률이 'N/A'로 표시됩니다.",
    "   - 계정과목명은 기업/업종에 따라 다르게 표기될 수 있어 자동 매칭을 사용했습니다.",
    "   - 배당을 실시하지 않았거나 공시에 해당 항목이 없는 경우 '-' 로 표시됩니다.",
    "   - 값은 원(KRW) 단위입니다.",
]
for i, line in enumerate(lines, start=2):
    c = ws0.cell(row=i, column=1, value=line)
    c.font = BOLD if (line and not line.startswith(" ")) else BLACK

# ============ 원본데이터 ============
ws1 = wb.create_sheet("원본데이터")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 20
for col in YEAR_COLS:
    ws1.column_dimensions[col].width = 20
ws1.merge_cells("A1:D1")
ws1["A1"] = "3개 기업 재무제표 원본 데이터 (단위: 원)"
ws1["A1"].font = TITLE_FONT

row_map = {}
r = 3
for company in COMPANIES:
    info = DATA.get(company, {})
    if "error" in info:
        ws1.cell(row=r, column=1, value=f"{company}: 데이터 조회 실패 ({info['error']})").font = Font(name=FONT_NAME, color="FF0000", bold=True)
        r += 2
        continue
    years = [info["bsns_year"] - 2, info["bsns_year"] - 1, info["bsns_year"]]
    ws1.merge_cells(f"A{r}:D{r}")
    c = ws1.cell(row=r, column=1, value=f"■ {company} ({info['dart_corp_name']})")
    c.font = BOLD_WHITE
    c.fill = COMPANY_FILL
    r += 1
    source_note = (f"출처: DART Open API, corp_code={info['corp_code']}, 종목코드={info['stock_code']}, "
                   f"사업연도={info['bsns_year']}(사업보고서), 재무제표구분={info['fs_div']} "
                   f"({'연결' if info['fs_div'] == 'CFS' else '별도'}), 배당: alotMatter 공시")
    ws1.merge_cells(f"A{r}:D{r}")
    ws1.cell(row=r, column=1, value=source_note).font = ITALIC_GRAY
    r += 1
    col_header_row = r
    ws1.cell(row=r, column=1, value="항목").font = BOLD_WHITE
    ws1.cell(row=r, column=1).fill = HEADER_FILL
    for col, yr in zip(YEAR_COLS, years):
        cell = ws1[f"{col}{r}"]
        cell.value = str(yr)
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    r += 1
    metric_rows = {"header": col_header_row, "years": years}
    for key, label, is_formula in METRICS:
        ws1.cell(row=r, column=1, value=label).font = BLACK
        if not is_formula:
            if key == "cash_dividend_total":
                periods = info.get("dividend", {}).get("cash_dividend_total", {})
            else:
                periods = info[key]
            for col, pk in zip(YEAR_COLS, PERIOD_KEYS):
                val = periods.get(pk)
                cell = ws1[f"{col}{r}"]
                cell.value = val if val is not None else "N/A"
                cell.font = BLUE
                cell.number_format = NUM_FMT
                cell.border = BORDER
        metric_rows[key] = r
        r += 1
    gp_row = metric_rows["gross_profit"]
    rev_row = metric_rows["revenue"]
    cogs_row = metric_rows["cogs"]
    for col in YEAR_COLS:
        cell = ws1[f"{col}{gp_row}"]
        cell.value = f"=IFERROR({col}{rev_row}-{col}{cogs_row},\"N/A\")"
        cell.font = BLACK
        cell.number_format = NUM_FMT
        cell.border = BORDER
    row_map[company] = metric_rows
    r += 1

ws1.freeze_panes = "B4"

# ============ 비교분석 ============
ws2 = wb.create_sheet("비교분석")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 16
for col in YEAR_COLS:
    ws2.column_dimensions[col].width = 14
ws2.merge_cells("A1:D1")
ws2["A1"] = "3개 기업 재무비율 비교분석"
ws2["A1"].font = TITLE_FONT

YEARS = row_map[COMPANIES[0]]["years"]
RAW = "원본데이터"

RATIO_SECTIONS = [
    ("[ 수익성 지표 ]", [
        ("매출총이익률", "gross_profit", "revenue", "0.0%"),
        ("영업이익률", "op_income", "revenue", "0.0%"),
        ("순이익률", "net_income", "revenue", "0.0%"),
        ("ROA (총자산순이익률)", "net_income", "total_assets", "0.0%"),
        ("ROE (자기자본순이익률)", "net_income", "total_equity", "0.0%"),
    ]),
    ("[ 안정성 지표 ]", [
        ("부채비율", "total_liab", "total_equity", "0.0%"),
        ("유동비율", "curr_assets", "curr_liab", "0.0%"),
        ("자기자본비율", "total_equity", "total_assets", "0.0%"),
    ]),
    ("[ 배당 지표 ]", [
        ("배당성향", "cash_dividend_total", "net_income", "0.0%"),
    ]),
]

ratio_row_map = {}

def write_ratio_table(start_row, section_title, ratio_defs):
    r = start_row
    ws2.merge_cells(f"A{r}:D{r}")
    ws2.cell(row=r, column=1, value=section_title).font = BOLD
    ws2.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    for label, num_key, den_key, fmt in ratio_defs:
        ws2.cell(row=r, column=1, value=label).font = BOLD_WHITE
        ws2.cell(row=r, column=1).fill = HEADER_FILL
        for col, yr in zip(YEAR_COLS, YEARS):
            c = ws2[f"{col}{r}"]
            c.value = str(yr)
            c.font = BOLD_WHITE
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        r += 1
        ratio_row_map[label] = {}
        for company in COMPANIES:
            rm = row_map[company]
            ws2.cell(row=r, column=1, value=company).font = BLACK
            for col in YEAR_COLS:
                cell = ws2[f"{col}{r}"]
                cell.value = f"=IFERROR('{RAW}'!{col}{rm[num_key]}/'{RAW}'!{col}{rm[den_key]},\"-\")"
                cell.number_format = fmt
                cell.font = BLACK
                cell.border = BORDER
            ratio_row_map[label][company] = r
            r += 1
        r += 1
    return r

row_cursor = 3
for section_title, defs in RATIO_SECTIONS:
    row_cursor = write_ratio_table(row_cursor, section_title, defs)

GROWTH_DEFS = [
    ("매출액증가율", "revenue"),
    ("영업이익증가율", "op_income"),
    ("순이익증가율", "net_income"),
    ("총자산증가율", "total_assets"),
]
r = row_cursor
ws2.merge_cells(f"A{r}:D{r}")
ws2.cell(row=r, column=1, value="[ 성장성 지표 (전년 대비 증가율) ]").font = BOLD
ws2.cell(row=r, column=1).fill = SECTION_FILL
r += 1
for label, key in GROWTH_DEFS:
    ws2.cell(row=r, column=1, value=label).font = BOLD_WHITE
    ws2.cell(row=r, column=1).fill = HEADER_FILL
    for col, yr in zip(YEAR_COLS, YEARS):
        c = ws2[f"{col}{r}"]
        c.value = str(yr)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    r += 1
    ratio_row_map[label] = {}
    for company in COMPANIES:
        rm = row_map[company]
        ws2.cell(row=r, column=1, value=company).font = BLACK
        ws2[f"B{r}"] = "N/A"
        ws2[f"B{r}"].alignment = Alignment(horizontal="center")
        ws2[f"C{r}"] = f"=IFERROR(('{RAW}'!C{rm[key]}-'{RAW}'!B{rm[key]})/'{RAW}'!B{rm[key]},\"-\")"
        ws2[f"D{r}"] = f"=IFERROR(('{RAW}'!D{rm[key]}-'{RAW}'!C{rm[key]})/'{RAW}'!C{rm[key]},\"-\")"
        for col in ["C", "D"]:
            ws2[f"{col}{r}"].number_format = "0.0%"
            ws2[f"{col}{r}"].font = BLACK
        for col in YEAR_COLS:
            ws2[f"{col}{r}"].border = BORDER
        ratio_row_map[label][company] = r
        r += 1
    r += 1
ws2.freeze_panes = "B3"

# ============ 대시보드 ============
ws3 = wb.create_sheet("대시보드")
ws3.sheet_view.showGridLines = False
for col, w in zip(["A", "B", "C", "D", "E", "F", "G", "H"], [16, 12, 12, 12, 4, 16, 12, 12]):
    ws3.column_dimensions[col].width = w
ws3.merge_cells("A1:H1")
ws3["A1"] = "3개 기업 재무비교 대시보드 (LG유플러스 / HD현대 / 셀트리온)"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A2:H2")
ws3["A2"] = f"기준: {YEARS[0]}~{YEARS[-1]} 사업보고서 (연결재무제표 기준, DART Open API)"
ws3["A2"].font = Font(name=FONT_NAME, italic=True, color="808080")

def add_line_chart(anchor, title, metric_key):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.height = 8
    chart.width = 16
    cat_row = row_map[COMPANIES[0]]["header"]
    cats = Reference(ws1, min_col=2, max_col=4, min_row=cat_row, max_row=cat_row)
    for company in COMPANIES:
        row = row_map[company][metric_key]
        data = Reference(ws1, min_col=2, max_col=4, min_row=row, max_row=row)
        series = Series(data, title=company)
        chart.series.append(series)
    chart.set_categories(cats)
    for s in chart.series:
        s.smooth = False
        s.marker.symbol = "circle"
    ws3.add_chart(chart, anchor)

def add_ratio_bar_chart(anchor, title, ratio_label):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.style = 10
    chart.height = 8
    chart.width = 16
    rows = ratio_row_map[ratio_label]
    header_row = list(rows.values())[0] - 1
    cats = Reference(ws2, min_col=2, max_col=4, min_row=header_row, max_row=header_row)
    for company in COMPANIES:
        row = rows[company]
        data = Reference(ws2, min_col=2, max_col=4, min_row=row, max_row=row)
        series = Series(data, title=company)
        chart.series.append(series)
    chart.set_categories(cats)
    ws3.add_chart(chart, anchor)

add_line_chart("A4", "매출액 추이 (원)", "revenue")
add_line_chart("A21", "영업이익 추이 (원)", "op_income")
add_line_chart("A38", "당기순이익 추이 (원)", "net_income")
add_ratio_bar_chart("J4", "ROE 비교 (%)", "ROE (자기자본순이익률)")
add_ratio_bar_chart("J21", "부채비율 비교 (%)", "부채비율")
add_ratio_bar_chart("J38", "배당성향 비교 (%)", "배당성향")

score_start = 55
ws3.merge_cells(f"A{score_start}:H{score_start}")
ws3.cell(row=score_start, column=1, value=f"■ {YEARS[-1]}년 요약 스코어카드").font = BOLD
ws3.cell(row=score_start, column=1).fill = SECTION_FILL

SCORE_METRICS = [
    ("영업이익률", "0.0%"), ("순이익률", "0.0%"), ("ROE (자기자본순이익률)", "0.0%"),
    ("부채비율", "0.0%"), ("유동비율", "0.0%"), ("배당성향", "0.0%"), ("매출액증가율", "0.0%"),
]
hdr_row = score_start + 1
ws3.cell(row=hdr_row, column=1, value="회사명").font = BOLD_WHITE
ws3.cell(row=hdr_row, column=1).fill = HEADER_FILL
for j, (label, fmt) in enumerate(SCORE_METRICS):
    cell = ws3.cell(row=hdr_row, column=2 + j, value=label)
    cell.font = BOLD_WHITE
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
ws3.row_dimensions[hdr_row].height = 30

first_data_row = hdr_row + 1
for i, company in enumerate(COMPANIES):
    r = first_data_row + i
    ws3.cell(row=r, column=1, value=company).font = BLACK
    for j, (label, fmt) in enumerate(SCORE_METRICS):
        src_row = ratio_row_map[label][company]
        cell = ws3.cell(row=r, column=2 + j)
        cell.value = f"='비교분석'!D{src_row}"
        cell.number_format = fmt
        cell.font = BLACK
        cell.border = BORDER

last_data_row = first_data_row + len(COMPANIES) - 1
for j in range(len(SCORE_METRICS)):
    col_letter = chr(ord("B") + j)
    rng = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
    rule = ColorScaleRule(start_type="min", start_color="F8696B",
                          mid_type="percentile", mid_value=50, mid_color="FFEB84",
                          end_type="max", end_color="63BE7B")
    ws3.conditional_formatting.add(rng, rule)

note_row = last_data_row + 2
ws3.merge_cells(f"A{note_row}:H{note_row}")
ws3.cell(row=note_row, column=1,
         value="※ 매출총이익률/배당성향이 '-' 또는 'N/A'로 표시된 경우 해당 계정이 공시에 없거나(예: LG유플러스 매출원가) 배당이 없는 경우입니다.").font = Font(name=FONT_NAME, italic=True, color="808080")

wb.save(os.path.join(HERE, "재무비교분석기3.xlsx"))

with open(os.path.join(HERE, "row_map3.json"), "w", encoding="utf-8") as f:
    json.dump(row_map, f, ensure_ascii=False, indent=2)
with open(os.path.join(HERE, "ratio_row_map3.json"), "w", encoding="utf-8") as f:
    json.dump(ratio_row_map, f, ensure_ascii=False, indent=2)

print("v3 workbook saved")
