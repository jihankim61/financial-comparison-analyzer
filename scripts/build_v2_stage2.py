# -*- coding: utf-8 -*-
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
FONT_NAME = "Arial"

BLACK = Font(name=FONT_NAME, color="000000")
BOLD = Font(name=FONT_NAME, bold=True)
BOLD_WHITE = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="203864")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

with open(os.path.join(HERE, "row_map2.json"), encoding="utf-8") as f:
    ROW_MAP = json.load(f)

COMPANIES = ["한국전력", "LG에너지솔루션", "HD현대"]
YEAR_COLS = ["B", "C", "D"]
YEARS = ROW_MAP[COMPANIES[0]]["years"]

wb = load_workbook(os.path.join(HERE, "재무비교분석기2_stage1.xlsx"))
RAW = "원본데이터"

ws2 = wb.create_sheet("비교분석")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 16
for col in YEAR_COLS:
    ws2.column_dimensions[col].width = 14

ws2.merge_cells("A1:D1")
ws2["A1"] = "3개 기업 재무비율 비교분석"
ws2["A1"].font = TITLE_FONT

RATIO_SECTIONS = [
    ("[ 수익성 지표 ]", [
        ("매출총이익률", "gross_profit", "revenue", "0.0%"),
        ("영업이익률", "op_income", "revenue", "0.0%"),
        ("순이익률", "net_income", "revenue", "0.0%"),
        ("ROA (총자산순이익률)", "net_income", "total_assets", "0.0%"),
        ("ROE (자기자본순이익률)", "net_income", "total_equity", "0.0%"),
    ]),
    ("[ 안정성 지표 ]", [
        ("부채비율", "total_liab", "total_equity", "0.0%"),
        ("유동비율", "curr_assets", "curr_liab", "0.0%"),
        ("자기자본비율", "total_equity", "total_assets", "0.0%"),
    ]),
    ("[ 배당 지표 ]", [
        ("배당성향", "cash_dividend_total", "net_income", "0.0%"),
    ]),
]

ratio_row_map = {}

def write_ratio_table(start_row, section_title, ratio_defs):
    r = start_row
    ws2.merge_cells(f"A{r}:D{r}")
    ws2.cell(row=r, column=1, value=section_title).font = BOLD
    ws2.cell(row=r, column=1).fill = SECTION_FILL
    r += 1
    for label, num_key, den_key, fmt in ratio_defs:
        ws2.cell(row=r, column=1, value=label).font = BOLD_WHITE
        ws2.cell(row=r, column=1).fill = HEADER_FILL
        for col, yr in zip(YEAR_COLS, YEARS):
            c = ws2[f"{col}{r}"]
            c.value = str(yr)
            c.font = BOLD_WHITE
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        r += 1
        ratio_row_map[label] = {}
        for company in COMPANIES:
            rm = ROW_MAP[company]
            ws2.cell(row=r, column=1, value=company).font = BLACK
            for col in YEAR_COLS:
                cell = ws2[f"{col}{r}"]
                cell.value = f"=IFERROR('{RAW}'!{col}{rm[num_key]}/'{RAW}'!{col}{rm[den_key]},\"-\")"
                cell.number_format = fmt
                cell.font = BLACK
                cell.border = BORDER
            ratio_row_map[label][company] = r
            r += 1
        r += 1
    return r

row_cursor = 3
for section_title, defs in RATIO_SECTIONS:
    row_cursor = write_ratio_table(row_cursor, section_title, defs)

GROWTH_DEFS = [
    ("매출액증가율", "revenue"),
    ("영업이익증가율", "op_income"),
    ("순이익증가율", "net_income"),
    ("총자산증가율", "total_assets"),
]

r = row_cursor
ws2.merge_cells(f"A{r}:D{r}")
ws2.cell(row=r, column=1, value="[ 성장성 지표 (전년 대비 증가율) ]").font = BOLD
ws2.cell(row=r, column=1).fill = SECTION_FILL
r += 1
for label, key in GROWTH_DEFS:
    ws2.cell(row=r, column=1, value=label).font = BOLD_WHITE
    ws2.cell(row=r, column=1).fill = HEADER_FILL
    for col, yr in zip(YEAR_COLS, YEARS):
        c = ws2[f"{col}{r}"]
        c.value = str(yr)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    r += 1
    ratio_row_map[label] = {}
    for company in COMPANIES:
        rm = ROW_MAP[company]
        ws2.cell(row=r, column=1, value=company).font = BLACK
        ws2[f"B{r}"] = "N/A"
        ws2[f"B{r}"].alignment = Alignment(horizontal="center")
        ws2[f"C{r}"] = f"=IFERROR(('{RAW}'!C{rm[key]}-'{RAW}'!B{rm[key]})/'{RAW}'!B{rm[key]},\"-\")"
        ws2[f"D{r}"] = f"=IFERROR(('{RAW}'!D{rm[key]}-'{RAW}'!C{rm[key]})/'{RAW}'!C{rm[key]},\"-\")"
        for col in ["C", "D"]:
            ws2[f"{col}{r}"].number_format = "0.0%"
            ws2[f"{col}{r}"].font = BLACK
        for col in YEAR_COLS:
            ws2[f"{col}{r}"].border = BORDER
        ratio_row_map[label][company] = r
        r += 1
    r += 1

ws2.freeze_panes = "B3"

with open(os.path.join(HERE, "ratio_row_map2.json"), "w", encoding="utf-8") as f:
    json.dump(ratio_row_map, f, ensure_ascii=False, indent=2)

wb.save(os.path.join(HERE, "재무비교분석기2_stage2.xlsx"))
print("stage2 v2 saved")
